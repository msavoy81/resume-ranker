#!/usr/bin/env python3
"""
Resume Ranker — CLI entry point + Excel formatter.

Scoring logic lives in scoring.py — do not add scoring code here.
To understand how candidates are ranked, read scoring.py.

Usage:
    python ranker.py candidates.csv ./resumes/ --jd job.txt
    python ranker.py candidates.csv ./resumes/ --jd job.txt --output ranked.xlsx --workers 10
"""

import argparse
import os
import sys
import textwrap
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

import anthropic
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _load_api_key() -> str:
    """Return the Anthropic API key, checking env var then a local .env file."""
    key = os.environ.get("ANTHROPIC_API_KEY", "")
    if key:
        return key
    env_file = Path(__file__).parent / ".env"
    if env_file.exists():
        for line in env_file.read_text().splitlines():
            line = line.strip()
            if line.startswith("ANTHROPIC_API_KEY="):
                key = line.split("=", 1)[1].strip().strip('"').strip("'")
                if key:
                    return key
    return ""

from scoring import (
    find_resume_file,
    process_candidate,
    sort_key,
    SYSTEM_TEMPLATE,
    extract_resume_text,  # re-exported so importers don't need to know about scoring.py
    _print_lock,
)


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

DARK_BLUE  = "1F3864"
WHITE      = "FFFFFF"
GREEN      = "C6EFCE"
YELLOW     = "FFEB9C"
RED        = "FFC7CE"
LIGHT_GREY = "F2F2F2"
FLAG_RED   = "FF0000"
FLAG_FILL  = "FFD7D7"

COLUMN_WIDTHS = {
    "Rank":                  6,
    "FLAG":                  13,
    "Candidate Name":        26,
    "Email":                 30,
    "Applied At":            16,
    "Stage":                 22,
    "Greenhouse Location":   22,
    "Stability Tier":        14,
    "NY Signal":             10,
    "JD Fit Composite":      14,
    "JD Fit Tier 1\n(LangGraph/Bedrock/AWS)":   22,
    "JD Fit Tier 2\n(MLOps/DevOps/AI Agents)":  24,
    "JD Fit Tier 3\n(General AI/ML)":            20,
    "Keywords Found":        34,
    "Stability Rationale":   42,
    "NY Rationale":          36,
    "Fit Rationale":         42,
    "Detected Location":     22,
    "Most Recent Role (yrs)": 18,
    "Recent Degree Year":    16,
    "Resume Found":          12,
}

FIT_SCORE_COLS = {
    "JD Fit Composite",
    "JD Fit Tier 1\n(LangGraph/Bedrock/AWS)",
    "JD Fit Tier 2\n(MLOps/DevOps/AI Agents)",
    "JD Fit Tier 3\n(General AI/ML)",
}

PREFERRED_COLUMNS = [
    "Rank", "FLAG", "Candidate Name", "Email", "Applied At", "Stage",
    "Greenhouse Location", "Stability Tier", "NY Signal",
    "JD Fit Composite",
    "JD Fit Tier 1\n(LangGraph/Bedrock/AWS)",
    "JD Fit Tier 2\n(MLOps/DevOps/AI Agents)",
    "JD Fit Tier 3\n(General AI/ML)",
    "Keywords Found",
    "Stability Rationale", "NY Rationale", "Fit Rationale",
    "Detected Location", "Most Recent Role (yrs)", "Recent Degree Year", "Resume Found",
]


def _stability_fill(tier: str) -> PatternFill:
    color = {"A": GREEN, "B": YELLOW, "C": RED}.get(tier)
    return (PatternFill(start_color=color, end_color=color, fill_type="solid")
            if color else PatternFill())


def _score_fill(value) -> PatternFill:
    try:
        v = float(value)
    except (TypeError, ValueError):
        return PatternFill()
    color = GREEN if v >= 7 else (YELLOW if v >= 4 else RED)
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def write_excel(df: pd.DataFrame, output_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ranked Candidates"

    header_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    header_font = Font(color=WHITE, bold=True, size=11)
    alt_fill    = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid")
    columns     = list(df.columns)

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        is_alt = (row_idx % 2 == 0)
        for col_idx, value in enumerate(row, 1):
            col_name = columns[col_idx - 1]
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            if col_name == "FLAG" and value:
                cell.fill = PatternFill(start_color=FLAG_FILL, end_color=FLAG_FILL, fill_type="solid")
                cell.font = Font(bold=True, color=FLAG_RED, size=10)
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            elif col_name == "Stability Tier" and value:
                cell.fill = _stability_fill(str(value))
                cell.font = Font(bold=True, size=12)
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif col_name == "NY Signal":
                color = GREEN if value == "Yes" else LIGHT_GREY
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif col_name in FIT_SCORE_COLS and value is not None:
                cell.fill = _score_fill(value)
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                if is_alt:
                    cell.fill = alt_fill
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col_idx, col_name in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS.get(col_name, 20)

    ws.freeze_panes = "C2"
    wb.save(output_path)


# ---------------------------------------------------------------------------
# Greenhouse column helpers
# ---------------------------------------------------------------------------

def _find_col(df: pd.DataFrame, candidates: list) -> "str | None":
    lower_map = {c.lower(): c for c in df.columns}
    for candidate in candidates:
        if candidate in df.columns:
            return candidate
        if candidate.lower() in lower_map:
            return lower_map[candidate.lower()]
    return None


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

DEFAULT_WORKERS = 8


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Rank Greenhouse candidates by stability, NY location, and JD fit.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=textwrap.dedent("""\
            Examples:
              python ranker.py export.csv ./resumes/ --jd job.txt
              python ranker.py export.csv ./resumes/ --jd job.txt --output ranked.xlsx --workers 10
        """),
    )
    parser.add_argument("csv",         help="Greenhouse CSV export")
    parser.add_argument("resumes_dir", help="Folder of resume files (PDF/DOCX/TXT)")
    parser.add_argument("--jd",        required=True, help="Job description text file")
    parser.add_argument("--output",    default="ranked_candidates.xlsx")
    parser.add_argument(
        "--workers", type=int, default=DEFAULT_WORKERS,
        help=f"Parallel Claude workers (default: {DEFAULT_WORKERS})",
    )
    args = parser.parse_args()

    csv_path    = Path(args.csv)
    resumes_dir = Path(args.resumes_dir)
    jd_path     = Path(args.jd)
    output_path = Path(args.output)

    for p, label in [(csv_path, "CSV"), (resumes_dir, "resumes dir"), (jd_path, "JD file")]:
        if not p.exists():
            print(f"Error: {label} not found: {p}", file=sys.stderr)
            sys.exit(1)

    api_key = _load_api_key()
    if not api_key:
        print("Error: Anthropic API key not found.", file=sys.stderr)
        print("  Set ANTHROPIC_API_KEY as an environment variable, or", file=sys.stderr)
        print("  create a .env file in this directory with:", file=sys.stderr)
        print("    ANTHROPIC_API_KEY=sk-ant-...", file=sys.stderr)
        sys.exit(1)
    os.environ["ANTHROPIC_API_KEY"] = api_key  # ensure subprocess/library picks it up

    print("Loading data...")
    df_gh = pd.read_csv(csv_path)

    name_col = _find_col(df_gh, ["Candidate Name", "Name", "Full Name", "candidate_name"])
    if name_col is None:
        print(f"Error: No name column found. Columns: {list(df_gh.columns)}", file=sys.stderr)
        sys.exit(1)

    today         = datetime.now().strftime("%Y-%m-%d")
    cutoff_year   = datetime.now().year - 3
    system_prompt = SYSTEM_TEMPLATE.format(cutoff_year=cutoff_year, today=today)

    resume_files = [
        f for f in resumes_dir.iterdir()
        if f.suffix.lower() in (".pdf", ".docx", ".doc", ".txt") and f.is_file()
    ]

    email_col    = _find_col(df_gh, ["Email", "Email Address"])
    stage_col    = _find_col(df_gh, ["Stage", "Current Stage", "Application Stage"])
    applied_col  = _find_col(df_gh, ["Applied At", "Application Date", "Date Applied"])
    location_col = _find_col(df_gh, ["Location", "Candidate Location"])

    total = len(df_gh)
    print(f"  {total} candidates in CSV")
    print(f"  {len(resume_files)} resume files in {resumes_dir}")
    print(f"  {args.workers} parallel workers  (keyword scoring: pure code, no API cost)\n")

    tasks: list = []
    for i, (_, row) in enumerate(df_gh.iterrows()):
        name        = str(row[name_col]).strip()
        resume_file = find_resume_file(name, resume_files)
        gh_row: dict = {}
        if email_col:    gh_row["Email"]               = row.get(email_col, "")
        if applied_col:  gh_row["Applied At"]          = row.get(applied_col, "")
        if stage_col:    gh_row["Stage"]               = row.get(stage_col, "")
        if location_col: gh_row["Greenhouse Location"] = row.get(location_col, "")
        tasks.append((i + 1, total, name, resume_file, gh_row))

    client = anthropic.Anthropic()

    results: list = [None] * total
    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        future_to_slot = {
            executor.submit(
                process_candidate,
                idx, total, client, name, resume_file, system_prompt, gh_row,
            ): (idx - 1)
            for idx, total, name, resume_file, gh_row in tasks
        }
        for future in as_completed(future_to_slot):
            slot = future_to_slot[future]
            try:
                results[slot] = future.result()
            except Exception as exc:
                name = tasks[slot][2]
                with _print_lock:
                    print(f"[ERROR] {name}: {exc}", flush=True)
                results[slot] = {
                    "_stability_tier": None, "_ny_signal": False, "_fit_composite": None,
                    "FLAG": "", "Candidate Name": name, "Stability Tier": "",
                    "NY Signal": "No", "JD Fit Composite": None, "Resume Found": "No",
                    "Stability Rationale": str(exc),
                }

    results.sort(key=sort_key)  # type: ignore[arg-type]

    out_df = pd.DataFrame(results)
    out_df.insert(0, "Rank", range(1, len(out_df) + 1))
    out_df = out_df.drop(columns=["_stability_tier", "_ny_signal", "_fit_composite"], errors="ignore")

    ordered = [c for c in PREFERRED_COLUMNS if c in out_df.columns]
    extras  = [c for c in out_df.columns if c not in ordered]
    out_df  = out_df[ordered + extras]

    print(f"\nWriting → {output_path}")
    write_excel(out_df, output_path)

    flagged   = (out_df.get("FLAG", pd.Series(dtype=str)) == "⚠ OPT REVIEW").sum()
    no_resume = (out_df.get("Resume Found", pd.Series(dtype=str)) == "No").sum()
    print(
        f"Done. {total - no_resume}/{total} candidates scored, "
        f"{flagged} flagged for OPT review, "
        f"{no_resume} skipped (no resume)."
    )


if __name__ == "__main__":
    main()
